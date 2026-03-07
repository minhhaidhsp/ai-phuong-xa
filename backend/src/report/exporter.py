"""
src/report/exporter.py
Sprint 6 - Xuat bao cao PDF va Excel
"""
import io
from datetime import date
from typing import Optional

from loguru import logger


# ── Excel Export (openpyxl) ───────────────────────────────────────────────────

def export_excel(bao_cao: dict) -> bytes:
    """
    Xuat bao cao thang thanh file .xlsx.
    Tra ve bytes de FastAPI tra ve FileResponse.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("Cai openpyxl: pip install openpyxl --break-system-packages")

    wb = Workbook()

    # Mau sac
    BLUE_FILL = PatternFill("solid", fgColor="1F4E79")
    LIGHT_FILL = PatternFill("solid", fgColor="D6E4F0")
    GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_cell(ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill = BLUE_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width

    def data_cell(ws, row, col, value, shade=False, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", size=10)
        c.fill = GRAY_FILL if shade else PatternFill("solid", fgColor="FFFFFF")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border
        if fmt:
            c.number_format = fmt

    # ── Sheet 1: Tong quan ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tổng Quan"
    ws1.row_dimensions[1].height = 30

    ws1["A1"] = f"BÁO CÁO THÁNG {bao_cao['thang']}/{bao_cao['nam']}"
    ws1["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Arial")
    ws1.merge_cells("A1:D1")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws1["A2"] = f"Từ ngày: {bao_cao['tu_ngay']}   |   Đến ngày: {bao_cao['den_ngay']}"
    ws1["A2"].font = Font(italic=True, size=10, color="555555", name="Arial")
    ws1.merge_cells("A2:D2")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # KPIs
    kpi_data = [
        ["Chỉ số", "Số lượng"],
        ["Tổng hồ sơ tiếp nhận", bao_cao["tong_tiep_nhan"]],
        ["Đã hoàn thành", bao_cao["tong_hoan_thanh"]],
        ["Từ chối", bao_cao["tong_tu_choi"]],
        ["Còn đang xử lý", bao_cao["con_xu_ly"]],
    ]
    ws1.row_dimensions[4].height = 25
    header_cell(ws1, 4, 1, "Chỉ số", width=30)
    header_cell(ws1, 4, 2, "Số lượng", width=15)
    for i, (k, v) in enumerate(kpi_data[1:], start=5):
        data_cell(ws1, i, 1, k, shade=(i % 2 == 0))
        data_cell(ws1, i, 2, v, shade=(i % 2 == 0), fmt="#,##0")

    # Theo linh vuc
    ws1["A10"] = "Theo lĩnh vực:"
    ws1["A10"].font = Font(bold=True, name="Arial", size=11, color="1F4E79")
    header_cell(ws1, 11, 1, "Lĩnh vực", width=30)
    header_cell(ws1, 11, 2, "Số hồ sơ", width=15)
    for i, lv in enumerate(bao_cao.get("theo_linh_vuc", []), start=12):
        data_cell(ws1, i, 1, lv["linh_vuc"], shade=(i % 2 == 0))
        data_cell(ws1, i, 2, lv["so_luong"], shade=(i % 2 == 0), fmt="#,##0")

    # ── Sheet 2: Hieu suat can bo ────────────────────────────────────────────
    ws2 = wb.create_sheet("Hiệu Suất Cán Bộ")
    ws2["A1"] = f"HIỆU SUẤT CÁN BỘ — Tháng {bao_cao['thang']}/{bao_cao['nam']}"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Arial")
    ws2.merge_cells("A1:E1")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    headers = ["Cán bộ", "Tổng hồ sơ", "Đã xong", "Đúng hạn", "Tỷ lệ đúng hạn (%)"]
    widths = [25, 15, 12, 12, 20]
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        header_cell(ws2, 3, j, h, width=w)

    for i, cb in enumerate(bao_cao.get("hieu_suat_can_bo", []), start=4):
        shade = i % 2 == 0
        data_cell(ws2, i, 1, cb["ho_ten"], shade)
        data_cell(ws2, i, 2, cb["so_ho_so"], shade, fmt="#,##0")
        data_cell(ws2, i, 3, cb["da_xong"], shade, fmt="#,##0")
        data_cell(ws2, i, 4, cb["dung_han"], shade, fmt="#,##0")
        c = ws2.cell(row=i, column=5, value=cb["ty_le_dung_han"])
        c.font = Font(name="Arial", size=10)
        c.fill = GRAY_FILL if shade else PatternFill("solid", fgColor="FFFFFF")
        c.alignment = Alignment(vertical="center")
        c.border = border
        c.number_format = '0.0"%"'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF Export (reportlab) ────────────────────────────────────────────────────

def export_pdf(bao_cao: dict) -> bytes:
    """
    Xuat bao cao thang thanh file .pdf.
    Dung reportlab SimpleDocTemplate.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise RuntimeError("Cai reportlab: pip install reportlab --break-system-packages")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    # Mau sac
    DARK_BLUE = colors.HexColor("#1F4E79")
    MID_BLUE = colors.HexColor("#2E75B6")
    LIGHT_BLUE = colors.HexColor("#D6E4F0")
    GRAY = colors.HexColor("#F2F2F2")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontName="Helvetica-Bold", fontSize=16,
        textColor=DARK_BLUE, spaceAfter=6, alignment=1
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontName="Helvetica", fontSize=10,
        textColor=colors.HexColor("#555555"),
        spaceAfter=12, alignment=1
    )
    heading_style = ParagraphStyle(
        "H2", parent=styles["Heading2"],
        fontName="Helvetica-Bold", fontSize=12,
        textColor=MID_BLUE, spaceBefore=12, spaceAfter=6
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"],
        fontName="Helvetica", fontSize=10, spaceAfter=4
    )

    def make_table(headers, rows, col_widths):
        data = [headers] + rows
        t = Table(data, colWidths=col_widths)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), DARK_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRAY]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWHEIGHT", (0, 0), (-1, -1), 20),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        t.setStyle(TableStyle(style_cmds))
        return t

    story = []

    # Tieu de
    story.append(Paragraph(
        f"BÁO CÁO THÁNG {bao_cao['thang']}/{bao_cao['nam']}",
        title_style
    ))
    story.append(Paragraph(
        f"HỆ THỐNG AI HÀNH CHÍNH PHƯỜNG/XÃ TP.HCM",
        sub_style
    ))
    story.append(Paragraph(
        f"Từ ngày {bao_cao['tu_ngay']} đến ngày {bao_cao['den_ngay']}",
        sub_style
    ))
    story.append(Spacer(1, 0.5*cm))

    # KPI tong quan
    story.append(Paragraph("I. CHỈ SỐ TỔNG QUAN", heading_style))
    kpi_rows = [
        ["Tổng hồ sơ tiếp nhận", str(bao_cao["tong_tiep_nhan"])],
        ["Đã hoàn thành", str(bao_cao["tong_hoan_thanh"])],
        ["Từ chối", str(bao_cao["tong_tu_choi"])],
        ["Còn đang xử lý", str(bao_cao["con_xu_ly"])],
    ]
    story.append(make_table(
        ["Chỉ số", "Số lượng"],
        kpi_rows,
        [12*cm, 4*cm]
    ))
    story.append(Spacer(1, 0.5*cm))

    # Theo linh vuc
    story.append(Paragraph("II. THEO LĨNH VỰC", heading_style))
    lv_rows = [
        [lv["linh_vuc"], str(lv["so_luong"])]
        for lv in bao_cao.get("theo_linh_vuc", [])
    ]
    if lv_rows:
        story.append(make_table(
            ["Lĩnh vực", "Số hồ sơ"],
            lv_rows,
            [12*cm, 4*cm]
        ))
    story.append(Spacer(1, 0.5*cm))

    # Hieu suat can bo
    story.append(Paragraph("III. HIỆU SUẤT CÁN BỘ", heading_style))
    cb_rows = [
        [
            cb["ho_ten"],
            str(cb["so_ho_so"]),
            str(cb["da_xong"]),
            str(cb["dung_han"]),
            f"{cb['ty_le_dung_han']}%"
        ]
        for cb in bao_cao.get("hieu_suat_can_bo", [])
    ]
    if cb_rows:
        story.append(make_table(
            ["Cán bộ", "Tổng", "Xong", "Đúng hạn", "Tỷ lệ"],
            cb_rows,
            [6*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]
        ))

    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(
        f"Ngày xuất báo cáo: {date.today()}  —  Hệ thống AI Hành chính Phường/Xã",
        body_style
    ))

    doc.build(story)
    return buf.getvalue()
